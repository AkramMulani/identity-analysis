#!/usr/bin/env python3
"""Bulk import persons from an Excel sheet into the facial recognition database.

The script mirrors the registration flow used by the Flask app:
- reads person details from an Excel workbook
- copies face images into the project photos directory
- stores photo records and face embeddings in the SQLite database
- copies fingerprint images to a temp folder and stores the extracted fingerprint template

Expected Excel columns (case-insensitive):
- SR_NO
- Name
- Age
- Address
- Add more face images
- Add more fingerprint images

Optional columns:
- Gender
- Notes
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Iterable, List, Optional, Sequence

ROOT_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT_DIR / "facial_project"
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled at runtime
    load_workbook = None  # type: ignore[assignment]

from backend import db, embedder
from backend.identity_analysis_utility import IdentityAnalysisUtility

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tif", ".tiff", ".webp"}


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def resolve_path(raw_path: Optional[object], base_dir: Path, candidates: Sequence[Path]) -> Optional[Path]:
    if raw_path is None:
        return None

    if isinstance(raw_path, (list, tuple)):
        for item in raw_path:
            resolved = resolve_path(item, base_dir, candidates)
            if resolved is not None:
                return resolved
        return None

    text = str(raw_path).strip()
    if not text:
        return None

    path = Path(text)
    if path.is_absolute():
        return path if path.exists() else None

    candidate_paths = [base_dir / path, Path.cwd() / path]
    for candidate in candidates:
        candidate_paths.append(candidate / path)

    for candidate in candidate_paths:
        if candidate.exists():
            return candidate.resolve()

    return None


def parse_excel_rows(excel_path: Path) -> Iterable[dict]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install it with: pip install openpyxl")

    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(cell) if cell is not None else "" for cell in rows[0]]
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        mapping = {}
        for index, header in enumerate(headers):
            if index < len(row):
                mapping[header] = row[index]
        if mapping:
            yield mapping


def split_multi_value(value: object) -> List[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        values = []
        for item in value:
            if item is not None:
                values.extend(split_multi_value(item))
        return values
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;,|\n]+", text) if part.strip()]


def get_value(mapping: dict, *possible_names: str) -> object:
    for name in possible_names:
        normalized_name = normalize_header(name)
        if normalized_name in mapping:
            return mapping[normalized_name]
    return None


def ensure_project_layout(project_root: Path) -> tuple[Path, Path, Path]:
    photos_dir = project_root / "photos"
    data_dir = project_root / "data"
    temp_dir = project_root / "temp"
    for directory in (photos_dir, data_dir, temp_dir):
        directory.mkdir(parents=True, exist_ok=True)
    return photos_dir, data_dir, temp_dir


def resolve_image_paths(raw_value: object, excel_path: Path, folder_name: str) -> List[Path]:
    values = split_multi_value(raw_value)
    resolved: List[Path] = []
    search_roots = [
        excel_path.parent / folder_name,
        excel_path.parent,
        ROOT_DIR / "Onboarding" / folder_name,
        ROOT_DIR / "Onboarding",
        PROJECT_ROOT,
        ROOT_DIR,
        ROOT_DIR / "Test",
        ROOT_DIR / "test",
    ]
    seen = set()

    for item in values:
        if not item:
            continue

        candidate = resolve_path(item, excel_path.parent, search_roots)
        if candidate is None:
            candidate = resolve_path(Path(str(item)).name, excel_path.parent, search_roots)

        if candidate is None:
            continue

        if candidate not in seen:
            seen.add(candidate)
            resolved.append(candidate)

    return resolved


def copy_image_to_photos(source_path: Path, photos_dir: Path, person_id: int, index: int) -> Path:
    suffix = source_path.suffix.lower() or ".jpg"
    if suffix not in IMAGE_EXTENSIONS:
        suffix = ".jpg"

    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", source_path.stem)[:40] or "image"
    destination = photos_dir / f"{person_id}_{index}_{stem}{suffix}"
    counter = 1
    while destination.exists():
        destination = photos_dir / f"{person_id}_{index}_{stem}_{counter}{suffix}"
        counter += 1

    shutil.copy2(source_path, destination)
    return destination


def import_person_row(row: dict, excel_path: Path, db_path: str, photos_dir: Path, temp_dir: Path, identity_utility: IdentityAnalysisUtility) -> int:
    name = str(get_value(row, "name") or "").strip()
    if not name:
        raise ValueError("Row is missing a person name")

    age_value = get_value(row, "age")
    age = int(age_value) if str(age_value).strip().isdigit() else None
    gender = str(get_value(row, "gender") or "").strip() or None
    address = str(get_value(row, "address") or "").strip() or None
    notes = str(get_value(row, "notes") or "").strip() or None

    person_id = db.add_person(db_path, name, age, gender, address, notes)

    face_images = resolve_image_paths(
        get_value(row, "addmorefaceimages", "faceimages", "images", "face image", "faceimages", "Facial_data"),
        excel_path,
        "Facial_data",
    )
    if not face_images:
        face_images = resolve_image_paths(get_value(row, "face"), excel_path, "Facial_data")

    if not face_images:
        print(f"No face images resolved for {name}; checked the Facial_data folder.")

    for index, image_path in enumerate(face_images, start=1):
        if not image_path.exists():
            print(f"Skipping missing image for {name}: {image_path}")
            continue

        destination = copy_image_to_photos(image_path, photos_dir, person_id, index)
        relative_path = destination.relative_to(PROJECT_ROOT).as_posix()
        photo_id = db.add_photo(db_path, person_id, relative_path)

        faces = embedder.extract_embeddings(str(image_path))
        for _, (emb, _, ratios) in enumerate(faces):
            db.add_embedding(db_path, person_id, photo_id, emb, ratios, model="insightface" if emb is not None else "fallback")

    fingerprint_images = resolve_image_paths(
        get_value(row, "addmorefingerprintimages", "fingerprintimages", "fingerprints", "fingerprint image", "fingerprintimages", "Fingerprint_data"),
        excel_path,
        "Fingerprint_data",
    )
    if not fingerprint_images:
        print(f"No fingerprint images resolved for {name}; checked the Fingerprint_data folder.")
    for fingerprint_path in fingerprint_images:
        if not fingerprint_path.exists():
            print(f"Skipping missing fingerprint image for {name}: {fingerprint_path}")
            continue

        temp_file = tempfile.NamedTemporaryFile(delete=False, dir=temp_dir, suffix=fingerprint_path.suffix or ".png")
        temp_path = temp_file.name
        temp_file.close()
        try:
            shutil.copy2(fingerprint_path, temp_path)
            fingerprint_template = identity_utility.extractFingerData(temp_path)
            if fingerprint_template:
                db.add_fingerprint(db_path, person_id, fingerprint_template)
        except Exception as exc:
            print(f"Warning: could not process fingerprint image for {name}: {exc}")
        finally:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except OSError:
                pass

    return person_id


def import_bulk(excel_path: Path, db_path: Optional[Path] = None) -> int:
    if db_path is None:
        db_path = PROJECT_ROOT / "data" / "facial.db"
    else:
        db_path = Path(db_path).expanduser().resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    photos_dir, _, temp_dir = ensure_project_layout(PROJECT_ROOT)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db.init_db(str(db_path))

    identity_utility = IdentityAnalysisUtility()
    imported_count = 0

    for row_number, row in enumerate(parse_excel_rows(excel_path), start=2):
        try:
            import_person_row(row, excel_path, str(db_path), photos_dir, temp_dir, identity_utility)
            imported_count += 1
            print(f"Imported person from row {row_number}")
        except Exception as exc:
            print(f"Skipping row {row_number}: {exc}")

    print(f"Import completed. Imported {imported_count} person(s) into {db_path}")
    return imported_count


def find_excel_file(explicit_path: Optional[str]) -> Path:
    if explicit_path:
        candidate = Path(explicit_path).expanduser()
        if not candidate.is_absolute():
            candidate = (Path.cwd() / candidate).resolve()
        return candidate

    for candidate in [ROOT_DIR / "Test" / "bulk_people.xlsx", ROOT_DIR / "Onboarding" / "bulk_people.xlsx", PROJECT_ROOT / "bulk_people.xlsx", ROOT_DIR / "bulk_people.xlsx"]:
        if candidate.exists():
            return candidate

    matches = sorted(ROOT_DIR.rglob("*.xlsx"))
    if matches:
        return matches[0]

    raise FileNotFoundError("No Excel file was found. Pass a workbook path explicitly or place one in the Test or Onboarding folder.")


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Bulk import persons from an Excel workbook into the facial database")
    parser.add_argument("excel_file", nargs="?", help="Path to the Excel file to import")
    parser.add_argument("--db-path", default=str(PROJECT_ROOT / "data" / "facial.db"), help="Target SQLite database path")
    return parser


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()

    excel_file = find_excel_file(args.excel_file)
    imported_count = import_bulk(excel_file, db_path=Path(args.db_path))
    return 0 if imported_count >= 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
